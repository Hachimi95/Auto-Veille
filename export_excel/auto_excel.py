import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import io
from datetime import datetime, date
import os
import json


def add_data_to_excel(workbook, sheet_name, input_data, client_config):
    """
    Add data to an Excel workbook and save it to the local file.
    """
    try:
        # Validate input_data is a dictionary
        if not isinstance(input_data, dict):
            raise ValueError("input_data must be a dictionary")

        # Get the sheet
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
        sheet = workbook[sheet_name]

        # Debug: Print sheet name and max row
        print(f"Processing sheet: {sheet.title}, Max row: {sheet.max_row}")

        # Load column mapping and formula columns
        column_mapping = client_config["column_mapping"]
        formula_columns = client_config.get("formula_columns", {})

        # Determine the next empty row
        next_row = sheet.max_row + 1 if sheet.max_row > 1 else 2
        print(f"Next empty row: {next_row}")

        # Write each field to its respective column
        for field, column in column_mapping.items():
            # Skip the "Client" field when iterating through input data
            if field == "Client":
                continue

            value = input_data.get(field, "")
            if isinstance(value, list):
                if field == "risques":
                    # Format with "-" separator between items (except first)
                    cleaned_items = []
                    for i, item in enumerate(value):
                        # Remove leading "-" if present
                        cleaned_item = item.strip()
                        if cleaned_item.startswith("-"):
                            cleaned_item = cleaned_item[1:].strip()
                        cleaned_items.append(cleaned_item)
                    
                    # Join with "-" separator between items
                    if len(cleaned_items) > 1:
                        value = cleaned_items[0]  # First item without prefix
                        for item in cleaned_items[1:]:
                            value += "\n-\n" + item
                    else:
                        value = "\n".join(cleaned_items)
                elif field == "Mitigations":
                    value = "\n".join(value) if isinstance(value, list) else value
                elif field in ["CVEs ID", "Références"]:
                    value = "\n".join(value)
            elif isinstance(value, str):
                # Handle string values that might contain comma-separated items
                if field == "risques":
                    # Split by comma and clean up
                    items = [item.strip() for item in value.split(",")]
                    cleaned_items = []
                    for item in items:
                        # Remove leading "-" if present
                        if item.startswith("-"):
                            item = item[1:].strip()
                        cleaned_items.append(item)
                    
                    # Join with "-" separator between items
                    if len(cleaned_items) > 1:
                        value = cleaned_items[0]  # First item without prefix
                        for item in cleaned_items[1:]:
                            value += "\n-\n" + item
                    else:
                        value = "\n".join(cleaned_items)
                elif field == "CVEs ID":
                    # Split by comma and clean up
                    items = [item.strip() for item in value.split(",")]
                    value = "\n".join(items)

            # Handle "Date de traitement" - use TODAY() function if it's today's date
            if field == "Date de traitement" and value:
                try:
                    # Try different date formats
                    date_formats = ["%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%m-%d-%Y", "%d-%m-%Y"]
                    date_obj = None
                    
                    for fmt in date_formats:
                        try:
                            date_obj = datetime.strptime(value, fmt)
                            break
                        except ValueError:
                            continue
                    
                    if not date_obj:
                        raise ValueError(f"Unable to parse date: {value}")
                    
                    # Check if the date is today
                    today = date.today()
                    is_today = (date_obj.date() == today)
                    
                    # If it's today's date, we'll handle it specially when setting the cell value
                    if not is_today:
                        # Store as a datetime object for non-today dates
                        value = date_obj
                    else:
                        # Mark for TODAY() formula handling
                        value = "TODAY"
                    
                except ValueError as e:
                    print(f"Error formatting date: {e}")
                    value = ""
            # Format "Date de notification" - convert to M/D/YYYY format
            elif field == "Date de notification" and value:
                try:
                    # Try different date formats
                    date_formats = ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y"]
                    date_obj = None
                    
                    for fmt in date_formats:
                        try:
                            date_obj = datetime.strptime(value, fmt)
                            break
                        except ValueError:
                            continue
                    
                    if date_obj:
                        # Convert to M/D/YYYY format
                        value = date_obj.strftime("%m/%d/%Y")
                    else:
                        print(f"Warning: Could not parse date '{value}' for Date de notification")
                        
                except Exception as e:
                    print(f"Error formatting Date de notification: {e}")
                    value = ""
            # Format other dates normally
            elif field == "Date" and value:
                try:
                    # Try different date formats
                    date_formats = ["%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%m-%d-%Y", "%d-%m-%Y"]
                    date_obj = None
                    
                    for fmt in date_formats:
                        try:
                            date_obj = datetime.strptime(value, fmt)
                            break
                        except ValueError:
                            continue
                    
                    if not date_obj:
                        raise ValueError(f"Unable to parse date: {value}")
                    
                    # Important: Store as a datetime object
                    value = date_obj
                    
                except ValueError as e:
                    print(f"Error formatting date: {e}")
                    value = ""

            # Cast "Delai" to a number if it's a valid digit
            if field == "Delai" and isinstance(value, str) and value.isdigit():
                value = int(value)

            # Set the cell value
            cell = sheet[f"{column}{next_row}"]
            
            # Special handling for Date de traitement when it's today's date
            if field == "Date de traitement":
                if value == "TODAY":
                    # Use Excel's TODAY() formula
                    cell.value = "=TODAY()"
                    # Ensure cell is formatted as a date
                    cell.number_format = "mm/dd/yyyy"
                elif isinstance(value, datetime):
                    if value.date() == date.today():
                        # Use Excel's TODAY() function for today's date
                        cell.value = "=TODAY()"
                        cell.number_format = "mm/dd/yyyy"
                    else:
                        # Regular date handling
                        cell.value = value
                        cell.number_format = "mm/dd/yyyy"
                else:
                    cell.value = value
            elif field == "Date" and isinstance(value, datetime):
                cell.value = value
                cell.number_format = "mm/dd/yyyy"
            else:
                cell.value = value
            
            # Style for "niveau risque"
            if field == "niveau risque":
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.font = Font(color="FFFFFF")

            # Style for "status"
            if field == "status":
                status_color = {
                    "Open": "FFFF00",      # Yellow
                    "OPEN": "FFFF00",      # Yellow
                    "WIP": "FFA500",       # Orange
                    "Pending": "87CEEB",   # Sky Blue
                    "NOK": "FF0000",       # Red
                    "Clos": "00B050",      # Green
                    "Clos (Traité)": "00B050",      # Green
                    "Clos (Patch cumulative)": "00B050",  # Green
                    "Clos (Non concerné)": "00B050"  # Green
                }.get(value, "FFFFFF")  # Default to white if unknown
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")

            # Add hyperlinks for "Références"
            if field == "Références":
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif field in ["Description", "Mitigations", "Remarque"]:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif field in ["CVEs ID", "risques"]:
                # Center align CVE IDs and Impact fields
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                # Apply alignment styles for all other fields
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write the client name to its specific column
        client_name = client_config.get("Client")
        client_column = column_mapping.get("Client", "A")  # Default to column A if not specified
        if client_name:
            client_cell = sheet[f"{client_column}{next_row}"]
            client_cell.value = client_name
            client_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add client-specific formulas and apply conditional coloring for "Deadline Status"
        for formula_field, formula_details in formula_columns.items():
            target_column = formula_details["column"]
            formula = formula_details["formula"].format(row=next_row)
            cell = sheet[f"{target_column}{next_row}"]
            cell.value = formula
            
            # Ensure formulas are properly recognized
            if formula.startswith('='):
                # This is crucial - set the data type explicitly for formulas
                cell.data_type = 'f'
                
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Apply conditional styling for "Deadline Status"
            if formula_field == "Deadline Status":
                calculated_value = eval_formula(input_data, column_mapping, formula_details, next_row)
                if calculated_value == "Traité dans le delai":
                    cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Autofit column width for better display
        for col in sheet.columns:
            max_length = 0
            col_letter = None  # Initialize column letter

            for cell in col:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue  # Skip merged cells

                if not col_letter:
                    col_letter = cell.column_letter  # Get column letter from first valid cell

                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            if col_letter:  # Ensure col_letter is not None
                adjusted_width = max(10, min(max_length + 2, 50))  # Limit column width
                sheet.column_dimensions[col_letter].width = adjusted_width

        # Resize the table and reapply the filter
        if sheet.tables:
            for table in sheet.tables.values():
                new_ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
                table.ref = new_ref
                print(f"Updated table {table.name} range to {new_ref}")

        if sheet.auto_filter.ref:
            sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

        print("Workbook modifications complete.")
        return workbook

    except Exception as e:
        print(f"An error occurred while modifying the workbook: {e}")
        raise


def eval_formula(input_data, column_mapping, formula_details, next_row):
    """
    Manually evaluate the formula based on input data to determine the condition's outcome.
    """
    start_date = input_data.get(column_mapping.get("Date"))
    end_date = input_data.get(column_mapping.get("Date de traitement"))
    delai = input_data.get(column_mapping.get("Delai"))

    if start_date and end_date and delai:
        try:
            date_diff = (int(end_date) - int(start_date)) if start_date.isdigit() and end_date.isdigit() else float('inf')
            return "Traité dans le delai" if date_diff <= int(delai) else "Hors délai de remediation"
        except Exception as e:
            print(f"Error evaluating formula: {e}")
    return "Hors délai de remediation"


def find_or_create_month_column(sheet, target_month, target_year):
    """
    Find the column for the given month/year in row 2, or create it if not present.
    Returns the column index.
    """
    for col in range(2, sheet.max_column + 1):
        cell = sheet.cell(row=2, column=col)
        val = cell.value
        if isinstance(val, datetime):
            if val.month == target_month and val.year == target_year:
                return col
        elif isinstance(val, str):
            try:
                # Try to parse as date string
                dt = datetime.strptime(val, "%m/%d/%Y")
                if dt.month == target_month and dt.year == target_year:
                    return col
            except Exception:
                # Try to match by month name (e.g., "June")
                try:
                    if datetime.strptime(val, "%B").month == target_month:
                        return col
                except Exception:
                    continue
    # If not found, add a new column at the end
    new_col = sheet.max_column + 1
    cell = sheet.cell(row=2, column=new_col)
    cell.value = datetime(target_year, target_month, 1)
    cell.number_format = "m/d/yyyy"
    return new_col

def update_recap_sheet(workbook, recap_sheet_name, status_counts, month, year):
    """
    Update the Récap sheet for the given month/year with status counts.
    status_counts: dict like {"Open": 3, "WIP": 2, ...}
    """
    sheet = workbook[recap_sheet_name]
    col = find_or_create_month_column(sheet, month, year)
    # Map status names to row indices (rows 8-13)
    status_row_map = {}
    for row in range(8, 14):
        status = sheet.cell(row=row, column=1).value
        if status:
            status_row_map[status.strip()] = row
    # Fill in the counts
    for status, count in status_counts.items():
        row = status_row_map.get(status)
        if row:
            sheet.cell(row=row, column=col).value = count
    return workbook


def _load_config():
    """Load client_config.json located in the same folder as this module."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(base_dir, 'client_config.json')
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _resolve_client_file_path(relative_path: str) -> str:
    """Resolve client Excel file path relative to this package directory."""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, relative_path)


def update_client_excel_file(client_name, data_to_add):
    """
    Update a client's Excel file with new data.
    
    Args:
        client_name (str): Name of the client (must match config key)
        data_to_add (list): List of dictionaries containing data to add
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Load client configuration (from package directory)
        config = _load_config()
        
        if client_name not in config:
            print(f"Client '{client_name}' not found in configuration")
            return False
        
        client_config = config[client_name]
        # Resolve file path relative to this package directory
        file_path = _resolve_client_file_path(client_config["file_path"])
        sheet_name = client_config["sheet"]
        
        # Check if file exists
        if not os.path.exists(file_path):
            print(f"❌ Client file not found: {file_path}")
            print(f"💡 Please ensure the Excel file exists at the specified path")
            return False
        
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        
        # Add each data row
        for data_row in data_to_add:
            workbook = add_data_to_excel(workbook, sheet_name, data_row, client_config)
        
        # Save the workbook
        try:
            workbook.save(file_path)
            print(f"Successfully updated {client_name} Excel file with {len(data_to_add)} new rows")
            return True
        except PermissionError as pe:
            print(f"Permission denied when saving {client_name} Excel file: {pe}")
            print(f"Please close the Excel file '{file_path}' and try again.")
            return False
        except Exception as save_error:
            print(f"Error saving {client_name} Excel file: {save_error}")
            return False
        
    except PermissionError as pe:
        print(f"Permission denied when loading {client_name} Excel file: {pe}")
        print(f"Please close the Excel file and try again.")
        return False
    except Exception as e:
        print(f"Error updating client Excel file: {e}")
        return False


def get_available_clients():
    """
    Get list of available clients from configuration.
    
    Returns:
        list: List of client names
    """
    try:
        config = _load_config()
        return list(config.keys())
    except Exception as e:
        print(f"Error loading client configuration: {e}")
        return []


