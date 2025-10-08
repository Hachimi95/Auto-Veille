import os
import re
import tempfile
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from collections import defaultdict
from datetime import datetime


def _extract_output_filename(input_filename: str) -> str:
    """Derive the output filename 'Mise à jour de sécurité <Month> <Year>.xlsx' from the input filename if possible."""
    date_patterns = [
        r'(\d{4})-(\d{2})-(\d{2})',  # YYYY-MM-DD
        r'(\d{2})/(\d{2})/(\d{4})',  # DD/MM/YYYY
        r'(\d{2})-(\d{2})-(\d{4})',  # DD-MM-YYYY
        r'(\d{4})_(\d{2})_(\d{2})',  # YYYY_MM_DD
        r'(\d{2})_(\d{2})_(\d{4})',  # DD_MM_YYYY
    ]

    extracted_date = None
    for pattern in date_patterns:
        match = re.search(pattern, input_filename)
        if match:
            groups = match.groups()
            if len(groups[0]) == 4:  # Year first
                year, month, day = groups
            else:  # Day first
                day, month, year = groups
            try:
                extracted_date = datetime(int(year), int(month), int(day))
                break
            except ValueError:
                continue

    if extracted_date:
        french_months = {
            1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
            5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
            9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'
        }
        month_name = french_months[extracted_date.month]
        year = extracted_date.year
        return f"Mise à jour de sécurité {month_name} {year}.xlsx"
    else:
        return f"Mise à jour de sécurité {datetime.now().strftime('%B %Y')}.xlsx"


def merge_excel_rows(input_file: str, output_file: str, sheet_name: str | None = None) -> None:
    """Merge rows per original logic into a new workbook and save to output_file."""
    wb = openpyxl.load_workbook(input_file)

    # Select the specified sheet or the active sheet if not specified
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
        sheet_name = ws.title

    # Get the header row
    header = [cell.value for cell in ws[1]]

    # Find the index of the "Article", "Download", and "Details" columns
    try:
        article_index = header.index("Article")
        download_index = header.index("Download")
        details_index = header.index("Details")
    except ValueError as e:
        raise Exception(f"Colonne manquante dans le fichier Excel: {e}")

    # Create a dictionary to store merged data
    merged_data = defaultdict(lambda: defaultdict(lambda: defaultdict(set)))

    # Iterate through rows and merge data
    for row in ws.iter_rows(min_row=2):
        article = row[article_index].value
        if not article:
            continue
        for i, cell in enumerate(row):
            if cell.value is not None:
                if i == article_index:
                    if not merged_data[article][i]:
                        merged_data[article][i] = (cell.value, cell.hyperlink)
                elif i == download_index:
                    merged_data[article][i]['text'].add(cell.value)
                    if cell.hyperlink:
                        merged_data[article][i]['link'].add(cell.hyperlink)
                elif i == details_index:
                    merged_data[article][i]['text'].add(cell.value)
                    if cell.hyperlink:
                        merged_data[article][i]['link'].add(cell.hyperlink)
                else:
                    merged_data[article][i]['data'].add(cell.value)

    # Create a new workbook for the merged data
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = f"{sheet_name}_Merged"

    # Write the header row
    for col, value in enumerate(header, start=1):
        new_ws.cell(row=1, column=col, value=value)

    # Define blue font for clickable cells
    blue_font = Font(color="0000FF", underline="single")

    # Write the merged data
    for row, (article, data) in enumerate(merged_data.items(), start=2):
        for col, values in data.items():
            cell = new_ws.cell(row=row, column=col + 1)
            if col == article_index:
                cell.value = values[0]
                if values[1]:
                    cell.hyperlink = values[1]
                    cell.font = blue_font
            elif col == download_index or col == details_index:
                cell.value = '\n'.join(sorted(set(values['text'])))
                if values['link']:
                    cell.hyperlink = next(iter(values['link']))
                    cell.font = blue_font
            else:
                cell.value = '\n'.join(sorted(set(str(v) for v in values['data'])))

            cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)

    # Auto-fit column widths with a maximum width
    max_width = 80
    for column in new_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min((max_length + 2), max_width)
        new_ws.column_dimensions[column_letter].width = adjusted_width

    # Adjust row heights
    for row in new_ws.rows:
        max_row_height = 0
        for cell in row:
            if cell.value:
                text_height = len(str(cell.value).split('\n')) * 15
                max_row_height = max(max_row_height, min(text_height, 250))
        new_ws.row_dimensions[row[0].row].height = max_row_height

    new_wb.save(output_file)


def process_uploaded_excel(file_storage, sheet_name: str | None = None):
    """
    Save the uploaded file to temp, merge rows, and return (output_path, output_filename, cleanup_fn).
    cleanup_fn() should be called after sending the response to delete temp files.
    """
    input_filename = file_storage.filename
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_input:
        file_storage.save(temp_input.name)
        input_path = temp_input.name

    output_filename = _extract_output_filename(input_filename)

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_output:
        output_path = temp_output.name

    merge_excel_rows(input_path, output_path, sheet_name)

    def cleanup_temp_files():
        try:
            if os.path.exists(input_path):
                os.remove(input_path)
            if os.path.exists(output_path):
                os.remove(output_path)
        except Exception:
            pass

    return output_path, output_filename, cleanup_temp_files


